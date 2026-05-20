from __future__ import annotations

import argparse
import json
import operator as op
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXPECTED_SHEETS = [
    "01_field_schema_33",
    "02a_mother_scoring",
    "02b_persona_rules",
    "02c_adjacent_resolver",
    "02d_scoring_constraints",
    "03_display_content",
]

EXPECTED_FIELD_COLUMNS = {
    "field_name": "field_key",
    "cn_name": "chinese_name",
    "range": "value_range",
    "source": "source",
    "confidence": "confidence_tier",
    "can_decide_persona": "can_decide_persona",
    "default_val": "default_value",
    "note": "notes",
}

FIELD_REQUIRED_COLUMNS = set(EXPECTED_FIELD_COLUMNS)
MOTHER_REQUIRED_COLUMNS = {
    "mother_id",
    "mother_name",
    "field",
    "operator",
    "weight",
    "score_formula",
    "field_level",
    "note",
}
PERSONA_REQUIRED_COLUMNS = {
    "persona_id",
    "persona_name",
    "mother_id",
    "field",
    "condition",
    "required",
    "weight",
    "note",
}
ADJACENT_REQUIRED_COLUMNS = {
    "persona_a",
    "persona_b",
    "field",
    "condition",
    "result",
}
DISPLAY_REQUIRED_COLUMNS = {
    "persona_id",
    "code",
    "persona_name",
    "hook",
    "tagline",
    "final_judgment",
    "poster_title",
    "poster_subtitle",
    "three_keywords",
    "share_copy",
}

COMPARATORS = {
    ">=": op.ge,
    "<=": op.le,
    ">": op.gt,
    "<": op.lt,
    "==": op.eq,
    "=": op.eq,
}


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def keyify(value: Any) -> str:
    return str(value or "").strip().upper()


def to_number(value: Any) -> Any:
    value = clean(value)
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value
    text = str(value).strip()
    try:
        number = float(text)
    except ValueError:
        return text
    return int(number) if number.is_integer() else number


def normalize_bool(value: Any) -> bool:
    text = str(value).strip().lower()
    return text in {"true", "yes", "1", "y"}


def normalize_confidence(value: Any) -> str:
    text = str(value or "").strip()
    mapping = {
        "核心": "core",
        "辅助": "auxiliary",
        "低置信": "low_confidence",
        "core": "core",
        "auxiliary": "auxiliary",
        "low_confidence": "low_confidence",
    }
    return mapping.get(text, text or "unknown")


def normalize_can_decide(value: Any) -> str:
    text = str(value or "").strip().lower()
    mapping = {
        "yes": "yes",
        "true": "yes",
        "1": "yes",
        "no": "no",
        "false": "no",
        "0": "no",
        "auxiliary": "auxiliary",
        "辅助": "auxiliary",
    }
    return mapping.get(text, "no")


def parse_condition_text(text: Any) -> dict[str, Any] | None:
    text = clean(text)
    if text is None:
        return None
    in_match = re.fullmatch(r"in\{([^}]+)\}", str(text).strip(), re.IGNORECASE)
    if in_match:
        values: list[Any] = []
        for part in in_match.group(1).split(","):
            number = float(part.strip())
            values.append(int(number) if number.is_integer() else number)
        return {"operator": "in", "threshold": values}
    match = re.fullmatch(r"(>=|<=|==|=|>|<)\s*(-?\d+(?:\.\d+)?)", str(text).strip())
    if not match:
        return None
    threshold = float(match.group(2))
    if threshold.is_integer():
        threshold = int(threshold)
    return {"operator": "==" if match.group(1) == "=" else match.group(1), "threshold": threshold}


def parse_inline_condition(text: str) -> dict[str, Any] | None:
    match = re.fullmatch(
        r"\s*([A-Z0-9_]+)\s*(>=|<=|==|=|>|<)\s*(-?\d+(?:\.\d+)?)\s*",
        text,
    )
    if not match:
        return None
    threshold = float(match.group(3))
    if threshold.is_integer():
        threshold = int(threshold)
    return {
        "field_key": match.group(1),
        "operator": "==" if match.group(2) == "=" else match.group(2),
        "threshold": threshold,
    }


def parse_adjacent_conditions(field_text: Any, condition_text: Any) -> tuple[str, str, Any, list[dict[str, Any]], str | None]:
    field_text = str(field_text or "").strip()
    condition = parse_condition_text(condition_text)

    if " AND " not in field_text and condition:
        return field_text, condition["operator"], condition["threshold"], [
            {
                "field_key": field_text,
                "operator": condition["operator"],
                "threshold": condition["threshold"],
            }
        ], None

    if condition_text is None and re.fullmatch(r"[A-Z0-9_]+", field_text):
        return field_text, "==", 1, [
            {
                "field_key": field_text,
                "operator": "==",
                "threshold": 1,
            }
        ], "Condition was blank in Excel; normalized to == 1 for binary field."

    parts = [part.strip() for part in field_text.split(" AND ")]
    parsed: list[dict[str, Any]] = []
    notes: list[str] = []
    for idx, part in enumerate(parts):
        inline = parse_inline_condition(part)
        if inline:
            parsed.append(inline)
            continue
        if idx == len(parts) - 1 and condition:
            parsed.append(
                {
                    "field_key": part,
                    "operator": condition["operator"],
                    "threshold": condition["threshold"],
                }
            )
            continue
        notes.append(f"Unable to parse adjacent condition segment: {part}")

    if parsed and not notes:
        return field_text, "compound", None, parsed, "Compound resolver normalized from split Excel columns."
    return field_text, "unparsed", None, parsed, "; ".join(notes) if notes else None


def rows_for_sheet(wb, sheet_name: str) -> tuple[list[str], list[dict[str, Any]], int]:
    ws = wb[sheet_name]
    headers = [clean(cell.value) for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    empty_rows = 0
    for values in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        normalized = [clean(value) for value in values[: len(headers)]]
        if not any(value is not None for value in normalized):
            empty_rows += 1
            continue
        rows.append(dict(zip(headers, normalized)))
    return [str(header) for header in headers if header], rows, empty_rows


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def build_field_schema(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    fields: list[dict[str, Any]] = []
    for row in rows:
        fields.append(
            {
                "field_key": keyify(row["field_name"]),
                "chinese_name": clean(row["cn_name"]),
                "value_range": clean(row["range"]),
                "source": clean(row["source"]),
                "confidence_tier": normalize_confidence(row["confidence"]),
                "can_decide_persona": normalize_can_decide(row["can_decide_persona"]),
                "default_value": to_number(row["default_val"]),
                "notes": clean(row["note"]),
            }
        )
    return fields


def build_mother_scoring(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rules: list[dict[str, Any]] = []
    for row in rows:
        field_key = keyify(row["field"])
        raw_operator = clean(row["operator"])
        score_formula = clean(row["score_formula"])
        operator_value = raw_operator
        threshold: Any = None
        if raw_operator is None and score_formula and "field==1" in score_formula.replace(" ", ""):
            operator_value = "if_eq_1_bonus"
            threshold = 1
        elif raw_operator == "in_set_2_3_bonus_15":
            threshold = [2, 3]
        elif raw_operator == "count":
            threshold = 2
        elif raw_operator == "compound_bonus_15":
            threshold = None

        confidence_tier = normalize_confidence(row["field_level"])
        rule = {
            "mother_id": clean(row["mother_id"]),
            "mother_name": clean(row["mother_name"]),
            "field_key": field_key,
            "operator": operator_value,
            "threshold": threshold,
            "weight": to_number(row["weight"]) or 0,
            "confidence_tier": confidence_tier,
            "is_core_support": confidence_tier == "core",
            "notes": clean(row["note"]),
        }
        if field_key == "[OTHER_M_SCORES>=60]":
            rule["special_rule"] = "other_m_scores_ge_60"
        elif "&" in field_key:
            rule["special_rule"] = "compound"
            rule["conditions"] = [
                {"field_key": "CHUAN_PALM", "operator": "==", "threshold": 1},
                {"field_key": "HEART_LINE_DEPTH", "operator": ">=", "threshold": 2},
            ]
        elif raw_operator == "in_set_2_3_bonus_15":
            rule["allowed_values"] = [2, 3]
        rules.append(rule)
    return rules


def build_persona_rules(rows: list[dict[str, Any]], todos: list[str]) -> list[dict[str, Any]]:
    rules: list[dict[str, Any]] = []
    for row in rows:
        parsed = parse_condition_text(row["condition"])
        if not parsed:
            if clean(row["condition"]) is None:
                operator_value = ">="
                threshold = 1
                normalized_note = "Blank Excel condition normalized to >= 1."
            else:
                todos.append(f"Persona rule condition could not be parsed: {row}")
                operator_value = "unparsed"
                threshold = None
                normalized_note = None
        else:
            operator_value = parsed["operator"]
            threshold = parsed["threshold"]
            normalized_note = None
        notes = clean(row["note"])
        if normalized_note:
            notes = f"{notes}; {normalized_note}" if notes else normalized_note
        rules.append(
            {
                "persona_id": clean(row["persona_id"]),
                "persona_name": clean(row["persona_name"]),
                "mother_id": clean(row["mother_id"]),
                "field_key": keyify(row["field"]),
                "operator": operator_value,
                "threshold": threshold,
                "weight": to_number(row["weight"]) or 0,
                "is_required": normalize_bool(row["required"]),
                "notes": notes,
            }
        )
    return rules


def build_adjacent_resolver(rows: list[dict[str, Any]], todos: list[str]) -> list[dict[str, Any]]:
    resolvers: list[dict[str, Any]] = []
    for row in rows:
        resolver_field, operator_value, threshold, conditions, note = parse_adjacent_conditions(row["field"], row["condition"])
        result_text = str(row["result"] or "").strip()
        match = re.fullmatch(r"(P\d{2});\s*else\s*(P\d{2})", result_text, re.IGNORECASE)
        if not match:
            todos.append(f"Adjacent resolver output could not be parsed: {row}")
            output_persona = None
            else_persona = None
        else:
            output_persona = match.group(1).upper()
            else_persona = match.group(2).upper()
        if operator_value in {"unparsed"}:
            todos.append(f"Adjacent resolver condition could not be fully parsed: {row}")
        resolvers.append(
            {
                "persona_a": clean(row["persona_a"]),
                "persona_b": clean(row["persona_b"]),
                "resolver_field": resolver_field,
                "operator": operator_value,
                "threshold": threshold,
                "output_persona": output_persona,
                "else_persona": else_persona,
                "conditions": conditions,
                "close_threshold_percent": 15,
                "notes": note,
            }
        )
    return resolvers


def build_scoring_constraints(rows: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "primary_mother_min_core_support": 2,
        "low_confidence_cannot_decide_primary_mother": True,
        "mother_score_cap": 100,
        "secondary_mother_gap": 15,
        "adjacent_close_threshold_percent": 15,
        "source_constraints": [
            {
                "constraint_id": clean(row["constraint_id"]),
                "description": clean(row["description"]),
                "condition": clean(row["condition"]),
                "action": clean(row["action"]),
                "scope": clean(row["scope"]),
            }
            for row in rows
        ],
    }


def build_display_content(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "persona_id": clean(row["persona_id"]),
            "persona_code": clean(row["code"]),
            "persona_name": clean(row["persona_name"]),
            "hook": clean(row["hook"]),
            "quote": clean(row["tagline"]),
            "final_judgement": clean(row["final_judgment"]),
            "poster_title": clean(row["poster_title"]),
            "poster_subtitle": clean(row["poster_subtitle"]),
            "three_keywords": clean(row["three_keywords"]),
            "share_copy": clean(row["share_copy"]),
        }
        for row in rows
    ]


def compare_condition(value: Any, operator_value: str, threshold: Any) -> bool:
    if value is None:
        value = 0
    func = COMPARATORS.get(operator_value)
    if func is None:
        return False
    return bool(func(value, threshold))


def inspect_and_validate(
    xlsx_path: Path,
    sheet_headers: dict[str, list[str]],
    sheet_rows: dict[str, list[dict[str, Any]]],
    empty_rows: dict[str, int],
    field_schema: list[dict[str, Any]],
    mother_scoring: list[dict[str, Any]],
    persona_rules: list[dict[str, Any]],
    adjacent_resolver: list[dict[str, Any]],
    display_content: list[dict[str, Any]],
    todos: list[str],
) -> str:
    source_note = f"Source workbook used: `{xlsx_path}`"
    field_keys = {field["field_key"] for field in field_schema}
    persona_ids = {f"P{i:02d}" for i in range(1, 37)}
    display_ids = {row["persona_id"] for row in display_content}
    mother_ids = {f"M{i}" for i in range(1, 9)}
    mother_present = {row["mother_id"] for row in mother_scoring}
    persona_present = {row["persona_id"] for row in persona_rules}
    persona_rule_counts = Counter(row["persona_id"] for row in persona_rules)

    missing_display = sorted(persona_ids - display_ids)
    missing_persona_rules = sorted(persona_ids - persona_present)
    too_few_rules = sorted(pid for pid in persona_ids if persona_rule_counts.get(pid, 0) < 2)
    missing_mothers = sorted(mother_ids - mother_present)

    for pid in missing_display:
        todos.append(f"Display content missing for {pid}.")
    for pid in missing_persona_rules:
        todos.append(f"Persona rules missing for {pid}.")
    for pid in too_few_rules:
        todos.append(f"Persona {pid} has fewer than 2 executable rules.")
    for mid in missing_mothers:
        todos.append(f"Mother type {mid} missing from mother scoring rules.")
    if len(field_schema) != 33:
        todos.append(f"Expected 33 palm fields, found {len(field_schema)}.")
    if len(adjacent_resolver) != 12:
        todos.append(f"Expected 12 adjacent resolver rows, found {len(adjacent_resolver)}.")

    referenced_fields: set[str] = set()
    unparseable: list[str] = []
    for rule in persona_rules:
        referenced_fields.add(rule["field_key"])
        if rule["operator"] == "unparsed":
            unparseable.append(f"Persona {rule['persona_id']} field {rule['field_key']}")
    for rule in mother_scoring:
        if rule.get("special_rule") in {"other_m_scores_ge_60", "compound"}:
            continue
        referenced_fields.add(rule["field_key"])
    for resolver in adjacent_resolver:
        for condition in resolver.get("conditions", []):
            referenced_fields.add(condition["field_key"])
        if resolver["operator"] == "unparsed":
            unparseable.append(f"Resolver {resolver['persona_a']}/{resolver['persona_b']}")

    unknown_fields = sorted(field for field in referenced_fields if field not in field_keys)
    for field in unknown_fields:
        todos.append(f"Referenced field not found in field schema: {field}")

    lines = [
        "# Excel Inspection Report",
        "",
        source_note,
        "",
        "## Sheet Summary",
        "",
        "| Sheet | Columns | Effective rows | Empty rows inside used range | Missing required columns |",
        "|---|---:|---:|---:|---|",
    ]
    required_by_sheet = {
        "01_field_schema_33": FIELD_REQUIRED_COLUMNS,
        "02a_mother_scoring": MOTHER_REQUIRED_COLUMNS,
        "02b_persona_rules": PERSONA_REQUIRED_COLUMNS,
        "02c_adjacent_resolver": ADJACENT_REQUIRED_COLUMNS,
        "02d_scoring_constraints": {"constraint_id", "description", "condition", "action", "scope"},
        "03_display_content": DISPLAY_REQUIRED_COLUMNS,
    }
    for sheet in EXPECTED_SHEETS:
        headers = sheet_headers.get(sheet, [])
        missing = sorted(required_by_sheet.get(sheet, set()) - set(headers))
        lines.append(
            f"| {sheet} | {len(headers)} | {len(sheet_rows.get(sheet, []))} | "
            f"{empty_rows.get(sheet, 0)} | {', '.join(missing) if missing else 'None'} |"
        )
    lines.extend(["", "## Sheet Columns", ""])
    for sheet in EXPECTED_SHEETS:
        lines.append(f"### {sheet}")
        lines.append(", ".join(sheet_headers.get(sheet, [])) or "Sheet not found")
        lines.append("")

    lines.extend(
        [
            "## Validation Checklist",
            "",
            f"- P01-P36 display copy complete: {'Yes' if not missing_display else 'No: ' + ', '.join(missing_display)}",
            f"- 33 palm fields complete: {'Yes' if len(field_schema) == 33 else 'No'}",
            f"- 8 mother types complete: {'Yes' if not missing_mothers else 'No: ' + ', '.join(missing_mothers)}",
            f"- 12 adjacent resolver rules complete: {'Yes' if len(adjacent_resolver) == 12 else 'No'}",
            f"- Unparseable executable rules: {'None' if not unparseable else ', '.join(unparseable)}",
            f"- Unknown referenced fields: {'None' if not unknown_fields else ', '.join(unknown_fields)}",
            "",
            "## Column Mapping",
            "",
        ]
    )
    for src, dst in EXPECTED_FIELD_COLUMNS.items():
        lines.append(f"- `01_field_schema_33.{src}` -> `{dst}`")
    lines.extend(
        [
            "- `02a_mother_scoring.field` -> `field_key`",
            "- `02b_persona_rules.condition` -> `operator` + `threshold`",
            "- `02c_adjacent_resolver.field` + `condition` -> `resolver_field` + `conditions`",
            "- `03_display_content.tagline` -> `quote`",
            "- `03_display_content.final_judgment` -> `final_judgement`",
            "",
            "## Notes",
            "",
            "- The workbook was not present in the project root; the user-provided file path was used.",
            "- Compound adjacent resolver rows were normalized into executable `conditions` arrays while preserving the required top-level fields.",
        ]
    )
    return "\n".join(lines) + "\n"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True)
    parser.add_argument("--project-root", default=str(Path(__file__).resolve().parents[1]))
    args = parser.parse_args()

    xlsx_path = Path(args.excel)
    project_root = Path(args.project_root)
    data_dir = project_root / "data"
    outputs_dir = project_root / "outputs"
    outputs_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx_path, data_only=True)
    todos: list[str] = []
    sheet_headers: dict[str, list[str]] = {}
    sheet_rows: dict[str, list[dict[str, Any]]] = {}
    empty_rows: dict[str, int] = {}

    missing_sheets = [sheet for sheet in EXPECTED_SHEETS if sheet not in wb.sheetnames]
    for sheet in missing_sheets:
        todos.append(f"Expected sheet missing: {sheet}")

    for sheet in EXPECTED_SHEETS:
        if sheet not in wb.sheetnames:
            sheet_headers[sheet] = []
            sheet_rows[sheet] = []
            empty_rows[sheet] = 0
            continue
        headers, rows, blanks = rows_for_sheet(wb, sheet)
        sheet_headers[sheet] = headers
        sheet_rows[sheet] = rows
        empty_rows[sheet] = blanks

    field_schema = build_field_schema(sheet_rows["01_field_schema_33"])
    mother_scoring = build_mother_scoring(sheet_rows["02a_mother_scoring"])
    persona_rules = build_persona_rules(sheet_rows["02b_persona_rules"], todos)
    adjacent_resolver = build_adjacent_resolver(sheet_rows["02c_adjacent_resolver"], todos)
    scoring_constraints = build_scoring_constraints(sheet_rows["02d_scoring_constraints"])
    display_content = build_display_content(sheet_rows["03_display_content"])

    report = inspect_and_validate(
        xlsx_path,
        sheet_headers,
        sheet_rows,
        empty_rows,
        field_schema,
        mother_scoring,
        persona_rules,
        adjacent_resolver,
        display_content,
        todos,
    )

    write_json(data_dir / "field_schema_33.json", field_schema)
    write_json(data_dir / "mother_scoring.json", mother_scoring)
    write_json(data_dir / "persona_rules.json", persona_rules)
    write_json(data_dir / "adjacent_resolver.json", adjacent_resolver)
    write_json(data_dir / "scoring_constraints.json", scoring_constraints)
    write_json(data_dir / "display_content.json", display_content)

    (outputs_dir / "excel_inspection_report.md").write_text(report, encoding="utf-8")
    todo_lines = ["# Validation Todo", ""]
    if todos:
        todo_lines.extend(f"- {item}" for item in dict.fromkeys(todos))
    else:
        todo_lines.append("- No blocking validation issues found in the Excel source.")
    todo_lines.append("")
    (outputs_dir / "validation_todo.md").write_text("\n".join(todo_lines), encoding="utf-8")
    print(f"Wrote normalized data and reports under {project_root}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
